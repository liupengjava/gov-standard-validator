#!/usr/bin/env python3
"""Spreadsheet parser for signal sample imports.

Outputs a JSON file path on stdout. The JSON contains {ok, text, sheets}.
"""
import argparse
import json
import os
import shutil
import subprocess
import tempfile


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def parse_xlsx(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    lines = []
    sheet_count = 0
    for sheet in workbook.worksheets:
        sheet_count += 1
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [clean(cell) for cell in row]
            while cells and not cells[-1]:
                cells.pop()
            if any(cells):
                rows.append(cells)
        if not rows:
            continue
        header = rows[0]
        data_rows = rows[1:] if len(rows) > 1 else rows
        for row in data_rows:
            parts = []
            for index, cell in enumerate(row):
                if not cell:
                    continue
                label = header[index] if index < len(header) and header[index] else f"列{index + 1}"
                parts.append(f"{label}：{cell}")
            if parts:
                lines.append("；".join(parts))
    return {"ok": True, "text": "\n".join(lines), "sheets": sheet_count, "parser": "openpyxl"}


def parse_xls(path):
    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError(f"当前运行环境缺少 pandas，无法解析 xls：{exc}")
    try:
        frames = pd.read_excel(path, sheet_name=None, dtype=str)
    except Exception as exc:
        tmpdir = tempfile.mkdtemp(prefix="gov_signal_xls_")
        try:
            subprocess.run(
                ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", tmpdir, path],
                check=True,
                capture_output=True,
                timeout=180,
            )
            converted = os.path.join(tmpdir, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
            if os.path.exists(converted):
                result = parse_xlsx(converted)
                result["parser"] = "libreoffice-openpyxl"
                return result
        except Exception:
            pass
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)
        raise RuntimeError(f"xls 解析失败，请另存为 xlsx/csv 后重试：{exc}")
    lines = []
    for _, frame in frames.items():
        frame = frame.fillna("")
        for _, row in frame.iterrows():
            parts = [f"{col}：{clean(value)}" for col, value in row.items() if clean(value)]
            if parts:
                lines.append("；".join(parts))
    return {"ok": True, "text": "\n".join(lines), "sheets": len(frames), "parser": "pandas"}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="input", required=True)
    parser.add_argument("--outdir", required=True)
    args = parser.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    ext = os.path.splitext(args.input)[1].lower()
    try:
        result = parse_xlsx(args.input) if ext == ".xlsx" else parse_xls(args.input)
    except Exception as exc:
        result = {"ok": False, "error": str(exc), "text": "", "sheets": 0}

    out = os.path.join(args.outdir, "sheet_parse_result.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)
    print(out)


if __name__ == "__main__":
    main()
